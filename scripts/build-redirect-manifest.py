import openpyxl, json, os, re, glob

REPO = "/Users/mitchel/Projects/untitled folder/prism"
XLSX = "/Users/mitchel/Desktop/PRISM website migration/prism-redirect-map.xlsx"

# Pages kept as real routes (destinations must be among these or homepage anchors).
KEEP = {"/", "/contact", "/history", "/are-current-llms-conscious",
        "/trustee-vacancies", "/blog", "/podcast", "/podcast-transcripts"}
RETIRED_BLOG = {
    "/blog/the-role-of-transparency-in-detecting-ai-consciousness",
    "/blog/prism-confronting-a-future-with-conscious-machines",
}
# Explicit reclassifications of former "Keep" pages → surviving section.
PEOPLE = ["/will-millership","/team","/mark-solms","/radhika-chadwick","/mitch-pass",
          "/calum-chace","/daniel-hulme","/ed-charvet","/nicholas-humphrey","/megan-peters",
          "/karl-friston","/irakli-beridze","/john-higgins","/susan-schneider"]
JOBS = ["/field-building-ops-coordinator","/researcher"]
EXPLICIT = {**{p: "/#people" for p in PEOPLE},
            **{p: "/#opportunities" for p in JOBS},
            "/principles": "/#values",
            "/the-field-of-artificial-consciousness": "/",  # retired: new map lives off-site
            **{p: "/blog" for p in RETIRED_BLOG}}

def is_kept(path):
    if path in KEEP: return True
    if path.startswith("/podcast/") or path.startswith("/podcast-transcripts/"): return True
    if path.startswith("/blog/") and path not in RETIRED_BLOG: return True
    return False

def classify(path):
    if path in EXPLICIT: return EXPLICIT[path]
    if path.startswith("/news"): return "/history"
    if path.startswith("/meetup"): return "/workshops"
    return "/"  # safe fallback

redirects = {}   # from -> to

# 1) The 24 rows already decided "Redirect" in the map.
wb = openpyxl.load_workbook(XLSX)
ws = wb["Redirect map"]
for r in list(ws.iter_rows(values_only=True))[5:]:
    if not r[0]: continue
    path, decision, dest = r[0], (r[5] or "").strip().lower(), (r[6] or "").strip()
    if decision.startswith("redirect") and dest:
        redirects[path] = dest

# 2) Former "Keep" rows we've reclassified to redirects.
for r in list(ws.iter_rows(values_only=True))[5:]:
    if not r[0]: continue
    path, decision = r[0], (r[5] or "").strip().lower()
    if decision.startswith("keep") and not is_kept(path):
        redirects[path] = classify(path)

# 3) Content-move redirects from front matter (highest precedence — specific target).
for md in glob.glob(os.path.join(REPO, "content", "*", "*.md")):
    coll = os.path.basename(os.path.dirname(md))
    prefix = {"podcast": "/podcast", "blog": "/blog"}.get(coll)
    if not prefix: continue
    txt = open(md).read()
    m = re.search(r"^---\n(.*?)\n---", txt, re.S)
    if not m: continue
    fm = m.group(1)
    if re.search(r"^\s*published:\s*false", fm, re.M): continue
    slug = re.sub(r"^\d{4}-\d{2}-\d{2}-", "", os.path.basename(md)[:-3])
    froms = re.findall(r"redirect_from:\s*\n((?:\s*-\s*\S+\n?)+)", fm)
    paths = []
    for block in froms:
        paths += re.findall(r"-\s*(\S+)", block)
    for f in paths:
        redirects[f] = f"{prefix}/{slug}"   # override any blanket rule

# 4) Resolve chains: follow /a -> /b -> /c to the final live target.
def resolve(dest, seen=None):
    seen = seen or set()
    while dest in redirects and dest not in seen:
        seen.add(dest); dest = redirects[dest]
    return dest
resolved = {frm: resolve(to) for frm, to in redirects.items()}
# Drop no-ops (a page redirecting to itself) and any that resolve to a kept page fine.
resolved = {k: v for k, v in resolved.items() if k != v}

out = dict(sorted(resolved.items()))
with open(os.path.join(REPO, "content", "redirects.json"), "w") as f:
    json.dump(out, f, indent=2)
    f.write("\n")

print(f"{len(out)} redirects written to content/redirects.json\n")
from collections import Counter
by_dest = Counter(out.values())
for dest, n in by_dest.most_common():
    print(f"  {n:3} → {dest}")
